# -*- coding: utf-8 -*-
"""
@File    : excel_utils.py
@Time    : 2026/7/12 15:55
@Author  : @叶风磊
@Desc    : 读取和操作excel文件的工具
"""

import openpyxl
from config.config import *

def read_excel(file_path = EXCEL_FILE,sheet_name = SHEET_NAME):
    #打开Excel 文件
    workbook = openpyxl.load_workbook(file_path) #参数文件名

    #选择表
    worksheet = workbook[sheet_name]

    #读取数据操作
    #思路：因为dict(zip(key,value)) 可以把读取到的数据变成字典类型，所以只需要分别取出key行，和value行即可
    #zip(worksheet) #zip（）函数的作用是可以把可迭代对象打包成一个个元组
    # print(dict(zip([1,2,3,4],[5,6,7])))
    data = [] #这是空列表，这是空列表

    keys = [cell.value for cell in worksheet[2]]      # 第1行：表头
    # values = [cell.value for cell in worksheet[3]]    # 第2行：数据
    for row in worksheet.iter_rows(min_row=3,values_only=True): #从第三行开始取数据，只返回值
        dict_data = dict(zip(keys, row))
        if dict_data["is_True"]:
            data.append(dict_data)
    # row_dict = dict(zip(keys, values))

    #关闭excel文件
    workbook.close()

    return data
print(read_excel())
# import openpyxl
# import os
# path = '../testdata.xlsx'
# print('exists:', os.path.exists(path))
# if os.path.exists(path):
#     wb = openpyxl.load_workbook(path)
#     ws = wb['Sheet1']
#     print('max_row:', ws.max_row, 'max_column:', ws.max_column)
#     print('row 1:', [c.value for c in ws[1]])
#     print('row 2 cells count:', len(list(ws[2])))
#     for cell in ws[2]:
#         print('cell:', cell, 'value:', cell.value)
#     wb.close()
